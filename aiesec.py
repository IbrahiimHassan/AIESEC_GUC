from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import smtplib
from email.message import EmailMessage
from email.utils import make_msgid

# Making the excel of yesterday up-to=date
import shutil
shutil.copy("Today.xlsx", "Yesterday.xlsx")
# Setup Chrome
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--user-data-dir=/tmp/chrome")

options.binary_location = "/usr/bin/chromium-browser"

service = Service("/usr/local/bin/chromedriver")

driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, 20)

# Step 1: Open the AIESEC GTA page
driver.get("https://aiesec.org/search?programmes=8")
# Accept cookie popup (always present)
cookie_btn = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//span[text()='Accept all cookies']/parent::button"))
)
try:
    cookie_btn.click()
except:
    driver.execute_script("arguments[0].click();", cookie_btn)

# Wait for the page to load after cookie click
time.sleep(30)
# Step 2: Click "Load more" until all cards are visible
while True:
     try:
         load_more_btn = wait.until(
             EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Load more']]"))
         )
         try:
             load_more_btn.click()
         except:
             driver.execute_script("arguments[0].click();", load_more_btn)

         time.sleep(30)  # Let new cards load
     except TimeoutException:
         break


# Step 3: Parse page
soup = BeautifulSoup(driver.page_source, "lxml")
driver.quit()

# Step 4: Extract opportunity info
card_data = {
    "OPPORTUNITY ID": [],
    "OPPORTUNITY LINK": [],
    "TITLE": [],
    "COUNTRY": [],
    "PREMIUM": [],
    "APPLICANTS": [],
    "DURATION": [],
    "ORGANIZATION": [],
    
}

for a in soup.find_all("a", href=True):
    href = a["href"]
    if "/opportunity/" not in href:
        continue

    full_link = "https://aiesec.org" + href
    opp_id = href.split("/")[-1]

    title_tag = a.find("h3")
    title = title_tag.get_text(strip=True) if title_tag else "N/A"

    premium = "Yes" if "Premium" in a.get_text() else "No"

    country = "N/A"
    duration = "N/A"
    duration_block = a.find("div", class_="flex flex-row items-center text-grey-dark text-[14px] flex-wrap")
    if duration_block:
        spans = duration_block.find_all("span")
        if spans and len(spans) >= 2:
            country = spans[0].get_text(strip=True)
            duration = spans[-1].get_text(strip=True)
            if duration == ".":
                duration = "N/A"


    applicants = "N/A"
    for div in a.find_all("div", class_="text-[12px]"):
        text = div.get_text(strip=True).lower()
        if "applicant" in text:
            applicants = text
            break

    org_block = a.find("div", class_="min-w-[80px]")
    organization = org_block.get_text(strip=True) if org_block else "N/A"

    card_data["OPPORTUNITY ID"].append(opp_id)
    card_data["OPPORTUNITY LINK"].append(full_link)
    card_data["TITLE"].append(title)
    card_data["COUNTRY"].append(country)
    card_data["PREMIUM"].append(premium)
    card_data["APPLICANTS"].append(applicants)
    card_data["DURATION"].append(duration)
    card_data["ORGANIZATION"].append(organization)
    


# Step 5: Save all data to Today.xlsx
df_today = pd.DataFrame(card_data)
df_today = df_today[["OPPORTUNITY ID", "OPPORTUNITY LINK", "TITLE", "COUNTRY", "PREMIUM", "APPLICANTS", "DURATION", "ORGANIZATION"]]
df_today.to_excel("Today.xlsx", index=False)
print(f"Total opportunities available today: {len(df_today)}")


# Step 6: Compare with Yesterday.xlsx to find new ones
df_yesterday = pd.read_excel("Yesterday.xlsx", sheet_name="Sheet1")

# Ensure both IDs are string for accurate comparison
df_today["OPPORTUNITY ID"] = df_today["OPPORTUNITY ID"].astype(str)
df_yesterday["OPPORTUNITY ID"] = df_yesterday["OPPORTUNITY ID"].astype(str)

new_df = df_today[~df_today["OPPORTUNITY ID"].isin(df_yesterday["OPPORTUNITY ID"])]
new_df = new_df[["OPPORTUNITY ID", "OPPORTUNITY LINK", "TITLE", "COUNTRY", "PREMIUM", "APPLICANTS", "DURATION", "ORGANIZATION"]]
new_df.to_excel("New.xlsx", index=False)

# Step 7: Format New.xlsx
wb = load_workbook("New.xlsx")
ws = wb.active

# Style headers
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

# Style Premium = Yes cells
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
premium_col_index = None
for i, col_name in enumerate(new_df.columns):
    if col_name == "PREMIUM":
        premium_col_index = i + 1
        break

if premium_col_index:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=premium_col_index)
        if cell.value == "Yes":
            cell.fill = yellow_fill

# Autofit column width
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 3

wb.save("New.xlsx")

# Step 8: Print result
print(f"Saved {len(new_df)} new opportunities to New.xlsx")

def generate_card_html(row):
    premium = row["PREMIUM"] == "Yes"
    premium_html = (
        '<div style="background-color: #FFD700; color: white; font-size: 0.65rem; padding: 4px 10px; '
        'border-radius: 20px; font-weight: bold; text-transform: uppercase;">PREMIUM</div>'
        if premium else ""
    )
    return f"""
    <div style="background-color: #ffffff; border-radius: 12px; box-shadow: 0 2px 6px rgba(0,0,0,0.1);
                padding: 16px; width: 300px; font-family: 'Segoe UI', sans-serif; display: inline-block;
                margin: 10px; vertical-align: top;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            {premium_html}
            <div style="font-size: 0.85rem; color: #6c63ff;">👥 {row['APPLICANTS']}</div>
        </div>
        <div style="font-weight: bold; font-size: 1.1rem; margin-top: 10px; color: #111;">{row['TITLE']}</div>
        <div style="color: #777; font-size: 0.9rem; margin-bottom: 6px;">{row['ORGANIZATION']}</div>
        <div style="font-size: 0.85rem; color: #444; line-height: 1.6;">
            🌍 {row['COUNTRY']}<br>
            ⏳ {row['DURATION']}
        </div>
        <div style="margin-top: 12px; border-top: 1px solid #eee; padding-top: 10px;">
            <a href="{row['OPPORTUNITY LINK']}" style="text-decoration: none; color: #1e3c72; font-weight: bold; font-size: 0.85rem;">
                Click to View Opportunity →
            </a>
        </div>
    </div>
    """


# Generate full email HTML
cards_html = "\n".join([generate_card_html(row) for _, row in new_df.iterrows()])

html_body = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>New AIESEC Opportunities</title>
  <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css" rel="stylesheet">
</head>
<body>
{cards_html}
</body>
</html>
"""

# Step 8 – Send the HTML email instead of attachment
if len(new_df) > 0:
    msg = EmailMessage()
    msg["Subject"] = "New AIESEC Opportunities Available"
    msg["From"] = "ogta.aiesecguc@gmail.com"
    msg["To"] = ["ibrahiim.hassan.04@gmail.com", "ahmed.sameh7433@gmail.com"]

    msg.set_content("There are new opportunities. Please view in HTML format.")
    msg.add_alternative(html_body, subtype='html')

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login("ogta.aiesecguc@gmail.com", "shsxmqvthntfvlrk")
        smtp.send_message(msg)

    print("Email with HTML cards sent successfully.")
else:
    print("No new data found. Email not sent.")
